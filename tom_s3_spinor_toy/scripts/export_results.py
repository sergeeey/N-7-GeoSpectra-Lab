"""Export gates_tracker.md to Excel and/or PDF.

Usage:
    python scripts/export_results.py --excel
    python scripts/export_results.py --pdf
    python scripts/export_results.py --all
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

# ── paths ────────────────────────────────────────────────────────────────────

ROOT = Path(__file__).resolve().parent.parent
TRACKER = ROOT / "docs" / "gates_tracker.md"
EXPORTS = ROOT / "docs" / "exports"


# ── markdown table parser ─────────────────────────────────────────────────────


def _parse_table(md_text: str, anchor: str = "<!-- Parser anchor") -> list[dict]:
    """Return rows from the first markdown table after `anchor` as list of dicts."""
    lines = md_text.splitlines()
    start = next((i for i, ln in enumerate(lines) if anchor in ln), None)
    if start is None:
        raise ValueError(f"Anchor '{anchor}' not found in tracker file.")

    header: list[str] = []
    rows: list[dict] = []
    in_table = False

    for line in lines[start:]:
        stripped = line.strip()
        if not stripped.startswith("|"):
            if in_table:
                break
            continue
        cells = [c.strip() for c in stripped.strip("|").split("|")]
        if all(re.fullmatch(r"[-: ]+", c) for c in cells):
            in_table = True
            continue
        if not in_table and not header:
            header = cells
            in_table = True  # next separator line will confirm
            continue
        if not header:
            header = cells
        else:
            rows.append(dict(zip(header, cells)))

    return rows


# ── status → colour mapping ───────────────────────────────────────────────────

STATUS_COLOURS = {
    "PROMOTE": "FF92D050",  # green
    "PASS": "FF00B050",  # dark green
    "STRONG_PASS": "FF00B050",
    "THEOREM": "FF0070C0",  # blue
    "PARTIAL": "FFFFEB9C",  # yellow
    "CONDITIONAL": "FFFFEB9C",
    "STRUCTURAL_SPLIT": "FFDCE6F1",  # light blue
    "NULL": "FFFF0000",  # red
    "WEAK": "FFFFB6B6",  # light red
    "OPEN": "FFFFC000",  # orange
}


# ── Excel export ──────────────────────────────────────────────────────────────


def export_excel(rows: list[dict], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Gates Tracker"

    if not rows:
        print("[export_excel] No rows to write.")
        return

    headers = list(rows[0].keys())

    # column widths (approx)
    col_widths = {
        "Gate": 12,
        "Section": 18,
        "Claim (one line)": 55,
        "Status": 14,
        "Key Result / Formula": 65,
        "Tests": 8,
        "Date": 12,
    }

    header_fill = PatternFill("solid", fgColor="FF203864")
    header_font = Font(bold=True, color="FFFFFFFF", size=10)
    thin = Side(border_style="thin", color="FFB8B8B8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # header row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = col_widths.get(h, 15)

    ws.row_dimensions[1].height = 22

    # data rows
    for row_idx, row in enumerate(rows, 2):
        status = row.get("Status", "")
        fill_colour = STATUS_COLOURS.get(status)
        fill = PatternFill("solid", fgColor=fill_colour) if fill_colour else None

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(h, ""))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            cell.font = Font(size=9)
            if fill and h == "Status":
                cell.fill = fill

        ws.row_dimensions[row_idx].height = 40

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[Excel] saved → {out_path}")


# ── PDF export ────────────────────────────────────────────────────────────────


def export_pdf(rows: list[dict], out_path: Path, tracker_path: Path) -> None:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)

    # read metadata from tracker header
    md_text = tracker_path.read_text(encoding="utf-8")
    last_updated = re.search(r"\*\*Last updated:\*\*\s*([^\|]+)", md_text)
    tests_line = re.search(r"\*\*Tests:\*\*\s*([^\n]+)", md_text)
    updated_str = last_updated.group(1).strip() if last_updated else "unknown"
    tests_str = tests_line.group(1).strip() if tests_line else "unknown"

    doc = SimpleDocTemplate(
        str(out_path),
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle(
        "cell",
        parent=styles["Normal"],
        fontSize=6.5,
        leading=8,
        wordWrap="LTR",
    )
    header_style = ParagraphStyle(
        "hdr",
        parent=styles["Normal"],
        fontSize=7,
        leading=9,
        textColor=colors.white,
        fontName="Helvetica-Bold",
    )
    title_style = ParagraphStyle(
        "title",
        parent=styles["Heading1"],
        fontSize=13,
        spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        "sub",
        parent=styles["Normal"],
        fontSize=8,
        textColor=colors.grey,
        spaceAfter=8,
    )

    STATUS_FILL = {
        "PROMOTE": colors.HexColor("#92D050"),
        "PASS": colors.HexColor("#00B050"),
        "STRONG_PASS": colors.HexColor("#00B050"),
        "THEOREM": colors.HexColor("#0070C0"),
        "PARTIAL": colors.HexColor("#FFEB9C"),
        "CONDITIONAL": colors.HexColor("#FFEB9C"),
        "STRUCTURAL_SPLIT": colors.HexColor("#DCE6F1"),
        "NULL": colors.HexColor("#FF9999"),
        "WEAK": colors.HexColor("#FFB6B6"),
        "OPEN": colors.HexColor("#FFC000"),
    }

    if not rows:
        print("[export_pdf] No rows.")
        return

    cols_to_show = ["Gate", "Section", "Claim (one line)", "Status", "Tests", "Date"]
    col_widths_pt = [30, 48, 165, 38, 22, 30]  # points; A4 landscape ≈ 277mm usable

    header_row = [Paragraph(h, header_style) for h in cols_to_show]
    table_data = [header_row]

    status_fills: dict[int, colors.HexColor] = {}

    for row_idx, row in enumerate(rows, 1):
        status = row.get("Status", "")
        if status in STATUS_FILL:
            status_fills[row_idx] = STATUS_FILL[status]
        table_data.append([Paragraph(row.get(h, ""), cell_style) for h in cols_to_show])

    t = Table(table_data, colWidths=col_widths_pt, repeatRows=1)

    base_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#203864")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 1), (-1, -1), 6.5),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#B8B8B8")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]

    # colour status column per row
    status_col = cols_to_show.index("Status")
    for row_idx, fill_color in status_fills.items():
        base_style.append(("BACKGROUND", (status_col, row_idx), (status_col, row_idx), fill_color))

    t.setStyle(TableStyle(base_style))

    story = [
        Paragraph("S³×S⁶ Research — Gates Tracker", title_style),
        Paragraph(
            f"Updated: {updated_str} · Tests: {tests_str} · "
            "Hard constraints: λ=FREE_COUPLING_PARAMETER · sm_derivation_claimed=False",
            sub_style,
        ),
        Spacer(1, 4 * mm),
        t,
    ]

    doc.build(story)
    print(f"[PDF]   saved → {out_path}")


# ── main ──────────────────────────────────────────────────────────────────────


def main() -> None:
    parser = argparse.ArgumentParser(description="Export gates_tracker.md to Excel/PDF")
    parser.add_argument("--excel", action="store_true")
    parser.add_argument("--pdf", action="store_true")
    parser.add_argument("--all", action="store_true")
    parser.add_argument("--tracker", default=str(TRACKER), help="Path to tracker MD file")
    args = parser.parse_args()

    if not (args.excel or args.pdf or args.all):
        parser.print_help()
        sys.exit(0)

    tracker_path = Path(args.tracker)
    if not tracker_path.exists():
        print(f"[ERROR] Tracker not found: {tracker_path}")
        sys.exit(1)

    md_text = tracker_path.read_text(encoding="utf-8")
    rows = _parse_table(md_text)
    print(f"[parse] {len(rows)} gates loaded from {tracker_path.name}")

    stem = tracker_path.stem

    if args.excel or args.all:
        export_excel(rows, EXPORTS / f"{stem}.xlsx")

    if args.pdf or args.all:
        export_pdf(rows, EXPORTS / f"{stem}.pdf", tracker_path)


if __name__ == "__main__":
    main()
